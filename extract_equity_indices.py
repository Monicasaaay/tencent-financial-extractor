import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import yfinance as yf
import warnings
import logging
import sys
from io import StringIO

warnings.filterwarnings('ignore')
# Suppress yfinance logging
logging.getLogger('yfinance').setLevel(logging.CRITICAL)
logging.disable(logging.CRITICAL)


class EquityIndicesExtractor:
    """
    Extract equity indices close prices for a date range and sort by region.
    Each date becomes a column with close prices.
    Automatically detects trading days vs non-trading days.
    """
    
    # Define indices by region with market and ticker information
    # Using reliable yfinance tickers with fallbacks
    INDICES_CONFIG = {
        'Asia': [
            {'market': 'Hong Kong', 'index_name': 'Hang Seng Index', 'ticker': '^HSI', 'alternatives': ['0001.HK']},
            {'market': 'Japan', 'index_name': 'Nikkei 225', 'ticker': '^N225', 'alternatives': ['998407.JP']},
            {'market': 'Singapore', 'index_name': 'Straits Times Index', 'ticker': '^STI', 'alternatives': []},
            {'market': 'South Korea', 'index_name': 'KOSPI', 'ticker': '^KS11', 'alternatives': []},
            {'market': 'Taiwan', 'index_name': 'Taiwan Weighted', 'ticker': '^TWII', 'alternatives': []},
            {'market': 'India', 'index_name': 'BSE SENSEX', 'ticker': '^BSESN', 'alternatives': []},
            {'market': 'Thailand', 'index_name': 'SET Index', 'ticker': '^SET.BK', 'alternatives': []},
        ],
        'Americas': [
            {'market': 'USA', 'index_name': 'S&P 500', 'ticker': '^GSPC', 'alternatives': []},
            {'market': 'USA', 'index_name': 'Nasdaq Composite', 'ticker': '^IXIC', 'alternatives': []},
            {'market': 'USA', 'index_name': 'Dow Jones Industrial', 'ticker': '^DJI', 'alternatives': []},
            {'market': 'Canada', 'index_name': 'S&P/TSX Composite', 'ticker': '^GSPTSE', 'alternatives': []},
            {'market': 'Mexico', 'index_name': 'Mexico IPC', 'ticker': '^MXX', 'alternatives': []},
            {'market': 'Brazil', 'index_name': 'Bovespa', 'ticker': '^BVSP', 'alternatives': []},
            {'market': 'Argentina', 'index_name': 'MERVAL', 'ticker': '^MERV', 'alternatives': []},
        ],
        'Europe': [
            {'market': 'UK', 'index_name': 'FTSE 100', 'ticker': '^FTSE', 'alternatives': []},
            {'market': 'Germany', 'index_name': 'DAX', 'ticker': '^GDAXI', 'alternatives': []},
            {'market': 'France', 'index_name': 'CAC 40', 'ticker': '^FCHI', 'alternatives': []},
            {'market': 'Spain', 'index_name': 'IBEX 35', 'ticker': '^IBEX', 'alternatives': []},
            {'market': 'Italy', 'index_name': 'FTSE MIB', 'ticker': 'FTSEMIB.MI', 'alternatives': []},
            {'market': 'Netherlands', 'index_name': 'AEX', 'ticker': '^AEX', 'alternatives': []},
            {'market': 'Switzerland', 'index_name': 'SMI', 'ticker': '^SSMI', 'alternatives': []},
            {'market': 'Sweden', 'index_name': 'OMX Stockholm 30', 'ticker': '^OMX', 'alternatives': []},
        ],
        'Middle East': [
            {'market': 'Saudi Arabia', 'index_name': 'TASI', 'ticker': '^TASI', 'alternatives': []},
            {'market': 'UAE', 'index_name': 'DFM General Index', 'ticker': '^DFMGI', 'alternatives': []},
            {'market': 'Qatar', 'index_name': 'Qatar Main Index', 'ticker': '^DSM', 'alternatives': []},
            {'market': 'Israel', 'index_name': 'TA-125', 'ticker': '^TA125.TA', 'alternatives': []},
        ],
        'Africa': [
            {'market': 'South Africa', 'index_name': 'JSE All Share', 'ticker': '^JALSH', 'alternatives': []},
            {'market': 'Egypt', 'index_name': 'EGX 30', 'ticker': '^EGX30.CA', 'alternatives': []},
            {'market': 'Nigeria', 'index_name': 'NSE All-Share', 'ticker': '^NSEINDX.LG', 'alternatives': []},
        ],
    }
    
    def __init__(self):
        """Initialize the extractor."""
        self.extracted_data = []
        self.failed_tickers = []
        self.trading_days = set()
        self.non_trading_days = set()
    
    def extract_for_date_range(self, start_date: str, end_date: str, 
                               regions: Optional[List[str]] = None,
                               use_alternatives: bool = True,
                               verbose: bool = False) -> pd.DataFrame:
        """
        Extract equity indices close prices for a date range.
        Each date becomes a column, with close prices as values.
        Automatically detects trading days.
        
        Args:
            start_date: Start date in 'YYYY-MM-DD' or 'DD-MMM-YYYY' format
            end_date: End date in 'YYYY-MM-DD' or 'DD-MMM-YYYY' format
            regions: List of regions to extract (None = all regions)
            use_alternatives: If True, try alternative tickers if primary fails
            verbose: If True, show verbose output including error messages
            
        Returns:
            DataFrame with equity indices sorted by region
        """
        # Suppress stderr temporarily
        old_stderr = sys.stderr
        sys.stderr = StringIO()
        
        try:
            # Parse dates
            start_obj = self._parse_date(start_date)
            end_obj = self._parse_date(end_date)
            
            start_formatted = start_obj.strftime('%Y-%m-%d')
            end_formatted = end_obj.strftime('%Y-%m-%d')
            
            print(f"\n{'='*70}")
            print(f"EQUITY INDICES EXTRACTION - DATE RANGE")
            print(f"{'='*70}")
            print(f"Start date: {start_formatted}")
            print(f"End date: {end_formatted}")
            print(f"Use alternatives for failed tickers: {use_alternatives}")
            print(f"Verbose mode: {verbose}\n")
            
            # Generate all dates in range
            all_dates = self._generate_date_range(start_obj, end_obj)
            print(f"Total dates in range: {len(all_dates)}\n")
            
            # Determine which regions to process
            regions_to_process = regions if regions else list(self.INDICES_CONFIG.keys())
            
            self.extracted_data = []
            self.failed_tickers = []
            self.trading_days = set()
            self.non_trading_days = set()
            
            total_indices = sum(
                len(self.INDICES_CONFIG[region]) 
                for region in regions_to_process 
                if region in self.INDICES_CONFIG
            )
            
            # Extract data for each region and index
            for region in regions_to_process:
                if region not in self.INDICES_CONFIG:
                    print(f"⚠ Warning: Region '{region}' not found. Skipping.")
                    continue
                
                print(f"📊 Processing {region}...")
                
                for index_config in self.INDICES_CONFIG[region]:
                    ticker = index_config['ticker']
                    market = index_config['market']
                    index_name = index_config['index_name']
                    alternatives = index_config.get('alternatives', [])
                    
                    # Fetch all historical data for this ticker
                    all_prices = self._fetch_prices_for_range(ticker, start_formatted, end_formatted, verbose)
                    
                    # If primary failed and alternatives available, try alternatives
                    if not all_prices and use_alternatives and alternatives:
                        for alt_ticker in alternatives:
                            all_prices = self._fetch_prices_for_range(alt_ticker, start_formatted, end_formatted, verbose)
                            if all_prices:
                                ticker = alt_ticker
                                break
                    
                    if all_prices:
                        # Build row with all dates
                        row_data = {
                            'Region': region,
                            'Market': market,
                            'Index Name': index_name,
                            'Ticker': ticker,
                        }
                        
                        # Add price for each date
                        for date_obj in all_dates:
                            date_str = date_obj.strftime('%Y-%m-%d')
                            date_col_name = date_obj.strftime('%d%b%Y')  # e.g., '01Jan2026'
                            
                            if date_str in all_prices:
                                # Trading day with data
                                row_data[date_col_name] = round(all_prices[date_str], 2)
                                self.trading_days.add(date_str)
                            else:
                                # Leave blank for non-trading days
                                row_data[date_col_name] = None
                        
                        self.extracted_data.append(row_data)
                        status = f"  ✓ {index_name}"
                    else:
                        self.failed_tickers.append(ticker)
                        status = f"  ✗ {index_name} ({ticker})"
                    
                    print(status)
            
            # Create DataFrame
            if self.extracted_data:
                df = pd.DataFrame(self.extracted_data)
                
                # Detect trading days by checking which dates have any non-null values
                date_columns = [col for col in df.columns if len(col) == 8 and col[2:3] not in ['g', 'a', 'k']]
                
                print(f"\n📅 Detecting trading days...")
                trading_days_detected = set()
                non_trading_days_detected = set()
                
                for col in date_columns:
                    if df[col].notna().any():
                        trading_days_detected.add(col)
                    else:
                        non_trading_days_detected.add(col)
                
                print(f"Trading days detected: {len(trading_days_detected)}")
                print(f"Non-trading days (blank): {len(non_trading_days_detected)}")
                
                # Sort by region, market, and index name
                df = df.sort_values(['Region', 'Market', 'Index Name']).reset_index(drop=True)
            else:
                df = pd.DataFrame()
            
            print()
            
            return df
        
        finally:
            # Restore stderr
            sys.stderr = old_stderr
    
    def _generate_date_range(self, start_obj: datetime, end_obj: datetime) -> List[datetime]:
        """Generate list of all dates between start and end (inclusive)."""
        dates = []
        current = start_obj
        while current <= end_obj:
            dates.append(current)
            current += timedelta(days=1)
        return dates
    
    def _parse_date(self, date_str: str) -> datetime:
        """Parse date string in multiple formats."""
        formats = [
            '%Y-%m-%d',      # 2026-06-04
            '%d-%b-%Y',      # 04-Jun-2026
            '%d-%m-%Y',      # 04-06-2026
            '%d/%m/%Y',      # 04/06/2026
            '%m/%d/%Y',      # 06/04/2026
            '%Y/%m/%d',      # 2026/06/04
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        raise ValueError(f"Date format not recognized: {date_str}. "
                        f"Use formats like: 2026-06-04 or 04-Jun-2026")
    
    def _fetch_prices_for_range(self, ticker: str, start_date: str, end_date: str, verbose: bool = False) -> Dict[str, float]:
        """
        Fetch close prices for a ticker across a date range.
        Completely suppresses all output from yfinance.
        
        Args:
            ticker: Stock ticker symbol
            start_date: Start date in 'YYYY-MM-DD' format
            end_date: End date in 'YYYY-MM-DD' format
            verbose: If True, show error messages
            
        Returns:
            Dictionary mapping date strings to close prices
        """
        try:
            # Suppress all output from yfinance
            old_stdout = sys.stdout
            old_stderr = sys.stderr
            sys.stdout = StringIO()
            sys.stderr = StringIO()
            
            try:
                # Download data for the range
                data = yf.download(ticker, start=start_date, end=end_date, progress=False)
                
                if data.empty or 'Close' not in data.columns:
                    return {}
                
                # Convert to dictionary: date -> close price
                prices = {}
                for date, row in data.iterrows():
                    close_val = row['Close']
                    if pd.notna(close_val):
                        date_str = date.strftime('%Y-%m-%d')
                        prices[date_str] = float(close_val)
                
                return prices
            
            finally:
                # Restore stdout/stderr
                sys.stdout = old_stdout
                sys.stderr = old_stderr
        
        except Exception as e:
            if verbose:
                print(f"    Error fetching {ticker}: {str(e)}")
            return {}
    
    def save_to_excel(self, df: pd.DataFrame, output_file: str):
        """Save extracted data to Excel file with formatting."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Equity Indices', index=False)
                worksheet = writer.sheets['Equity Indices']
                
                # Define formatting
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format header row
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Format data rows
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    row_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") \
                        if row_idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    for col_idx, cell in enumerate(row, start=1):
                        cell.border = thin_border
                        cell.fill = row_fill
                        
                        if col_idx <= 4:  # Region, Market, Index Name, Ticker
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # Price columns
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                
                # Adjust column widths
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 18
                worksheet.column_dimensions['C'].width = 30
                worksheet.column_dimensions['D'].width = 15
                
                # Set width for date columns
                for col_idx in range(5, worksheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 12
                
                worksheet.freeze_panes = 'A2'
            
            print(f"✓ Data saved to {output_file}")
            return True
        
        except ImportError:
            print("⚠ openpyxl not installed. Saving as CSV instead...")
            df.to_csv(output_file.replace('.xlsx', '.csv'), index=False)
            return False
    
    def save_to_csv(self, df: pd.DataFrame, output_file: str):
        """Save extracted data to CSV file."""
        df.to_csv(output_file, index=False)
        print(f"✓ Data saved to {output_file}")
    
    def get_available_regions(self) -> List[str]:
        """Get list of available regions."""
        return list(self.INDICES_CONFIG.keys())
    
    def get_region_config(self, region: str) -> List[Dict]:
        """Get configuration for a specific region."""
        return self.INDICES_CONFIG.get(region, [])


# Example usage
if __name__ == "__main__":
    extractor = EquityIndicesExtractor()
    
    print("\n" + "="*70)
    print("AVAILABLE REGIONS")
    print("="*70)
    for region in extractor.get_available_regions():
        count = len(extractor.get_region_config(region))
        print(f"  • {region}: {count} indices")
    print("="*70)
    
    # ============================================================
    # Extract for 01-Jan-2026 to 06-Jun-2026
    # ============================================================
    start_date = '01-Jan-2026'
    end_date = '06-Jun-2026'
    
    df = extractor.extract_for_date_range(start_date, end_date, use_alternatives=True, verbose=False)
    
    if not df.empty:
        print("EXTRACTED EQUITY INDICES DATA:")
        print("-" * 70)
        print(df.to_string(index=False))
        print()
        
        # Save to Excel and CSV
        extractor.save_to_excel(df, f'equity_indices_{start_date.replace("-", "")}_{end_date.replace("-", "")}.xlsx')
        extractor.save_to_csv(df, f'equity_indices_{start_date.replace("-", "")}_{end_date.replace("-", "")}.csv')
    else:
        print("⚠ No data extracted. Some tickers may not have historical data in yfinance.")
        print("Try with verbose=True to see detailed error messages.")
    
    # ============================================================
    # Extract for specific regions only
    # ============================================================
    print("\n" + "="*70)
    print("EXTRACTING ASIA & EUROPE REGIONS ONLY")
    print("="*70)
    
    df_filtered = extractor.extract_for_date_range(
        start_date,
        end_date,
        regions=['Asia', 'Europe'],
        use_alternatives=True,
        verbose=False
    )
    
    if not df_filtered.empty:
        print("FILTERED DATA (Asia & Europe):")
        print("-" * 70)
        print(df_filtered.to_string(index=False))
        print()
        
        # Save filtered data
        extractor.save_to_excel(df_filtered, f'equity_indices_asia_europe_{start_date.replace("-", "")}_{end_date.replace("-", "")}.xlsx')
    else:
        print("⚠ No data extracted for these regions.")
