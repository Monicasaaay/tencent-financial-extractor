import pdfplumber
import pandas as pd
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

# ==================== PAGE DETECTION UTILITY ====================

def find_financial_statement_pages(pdf_file, start_keyword, stop_keywords=None, unit_hint="RMB", min_numeric_lines=3):
    """
    Dynamically find page indices containing a financial statement table.
    Returns a list of page indices most likely to contain the desired statement.
    Raises an error if pages cannot be found.
    """
    with pdfplumber.open(pdf_file) as pdf:
        candidates = []
        currently_in_table = False
        stop_keywords = stop_keywords or []
        
        for idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            lines = [ln.strip() for ln in text.split('\n') if ln.strip()]
            
            # Look for start_keyword in the header (first few lines)
            found_header = any(start_keyword.lower() in line.lower() for line in lines[:5])
            
            # Check for stop keywords to know when table ends
            found_stop = any(any(stop_kw.lower() in line.lower() for stop_kw in stop_keywords) for line in lines[:5])
            
            if found_header:
                currently_in_table = True
            
            if currently_in_table and found_stop:
                break
            
            # Look for sufficient numeric lines and unit (RMB)
            numeric_lines = sum(
                bool(re.search(r'[0-9,\(\)]{3,}', ln)) for ln in lines
            )
            is_table_like = unit_hint in text and numeric_lines >= min_numeric_lines
            
            if currently_in_table and is_table_like:
                candidates.append(idx)
        
        return candidates


# ==================== EXTRACTION FUNCTIONS ====================

def convert_to_number(value_str):
    """
    Convert string number to float, handling various formats.
    Handles: "123,456" -> 123456, "(123,456)" -> -123456, "24.749" -> 24.749
    """
    if not value_str or value_str in ['–', '-', 'None', '—', '\\u2013', '\\u2014']:
        return None
    
    # Check if negative (in parentheses)
    is_negative = value_str.startswith('(') and value_str.endswith(')')
    
    # Remove parentheses, commas, and dashes
    clean_value = value_str.replace('(', '').replace(')', '').replace(',', '').replace('–', '').replace('—', '')
    
    try:
        num = float(clean_value)
        if is_negative:
            num = -num
        return num
    except ValueError:
        return None


def extract_income_statement(pdf_file, page_index):
    """
    Extract Consolidated Income Statement with note index extraction
    """
    with pdfplumber.open(pdf_file) as pdf:
        page = pdf.pages[page_index]
        page_text = page.extract_text()
    
    lines = page_text.split('\n')
    data_rows = []
    i = 0
    pending_note = None
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Skip headers and metadata
        if any(skip in line for skip in [
            'Consolidated Income Statement',
            'For the year ended',
            'Year ended 31 December',
            '2025 2024',
            'Note RMB',
            'The notes on pages',
            'Annual Report',
            'Earnings per share for profit attributable'
        ]):
            i += 1
            continue
        
        # Skip empty lines and section headers
        if not line or line in ['Revenues', 'Attributable to:', 'of the Company (RMB per share)']:
            i += 1
            continue
        
        # Check if this line is ONLY a note number
        if re.match(r'^\d+$', line) and len(line) <= 3:
            pending_note = line
            i += 1
            continue
        
        # Extract numbers from the line
        numbers = re.findall(r'\(?\d+(?:,\d{3})*(?:\.\d{2,3})?\)?', line)
        
        # Check if this line has financial data
        if len(numbers) >= 2:
            val_2024_str = numbers[-1]
            val_2025_str = numbers[-2]
            
            val_2024 = convert_to_number(val_2024_str)
            val_2025 = convert_to_number(val_2025_str)
            
            # Remove the numbers from the line to find item name
            line_text = re.sub(r'\(?\d+(?:,\d{3})*(?:\.\d{2,3})?\)?', '', line).strip()
            parts = line_text.split()
            
            note_index = pending_note if pending_note else ''
            item_name = line_text
            pending_note = None
            
            # Check if last part is a note reference
            if parts:
                last_part = parts[-1]
                if re.match(r'^\d+(?:\(.[\)])?$', last_part):
                    note_index = last_part
                    item_name = ' '.join(parts[:-1]).strip()
            
            data_rows.append({
                'Line item name': item_name,
                'Note index': note_index,
                'figure for 2025': val_2025,
                'figure for 2024': val_2024
            })
        else:
            # No financial data on this line
            if line and len(line) > 2:
                note_to_use = pending_note if pending_note else ''
                data_rows.append({
                    'Line item name': line,
                    'Note index': note_to_use,
                    'figure for 2025': None,
                    'figure for 2024': None
                })
                pending_note = None
        
        i += 1
    
    return pd.DataFrame(data_rows)


def extract_comprehensive_income_statement(pdf_file, page_index):
    """
    Extract Consolidated Statement of Comprehensive Income
    """
    with pdfplumber.open(pdf_file) as pdf:
        page = pdf.pages[page_index]
        page_text = page.extract_text()
    
    lines = page_text.split('\n')
    data_rows = []
    i = 0
    current_item = ""
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Skip headers and metadata
        if any(skip in line for skip in [
            'Consolidated Statement of Comprehensive Income',
            'For the year ended',
            'Year ended 31 December',
            '2025 2024',
            'RMB',
            'The notes on pages',
            'Tencent Holdings Limited'
        ]):
            i += 1
            continue
        
        if not line:
            i += 1
            continue
        
        # Extract numbers from the line
        numbers = re.findall(r'\(?\d+(?:,\d{3})*(?:\.\d{2,3})?\)?', line)
        
        # Check if this line has financial data
        if len(numbers) >= 2:
            val_2024_str = numbers[-1]
            val_2025_str = numbers[-2]
            
            val_2024 = convert_to_number(val_2024_str)
            val_2025 = convert_to_number(val_2025_str)
            
            # Remove numbers to get item name
            line_text = re.sub(r'\(?\d+(?:,\d{3})*(?:\.\d{2,3})?\)?', '', line).strip()
            
            # Combine with current_item if it was a multi-line item
            if current_item:
                item_name = (current_item + " " + line_text).strip()
                current_item = ""
            else:
                item_name = line_text
            
            data_rows.append({
                'Line item name': item_name,
                'Note index': '',
                'figure for 2025': val_2025,
                'figure for 2024': val_2024
            })
        else:
            # No financial data on this line
            if line in ['Other comprehensive income, net of tax:',
                       'Items that may be subsequently reclassified to profit or loss',
                       'Items that will not be subsequently reclassified to profit or loss',
                       'Attributable to:']:
                data_rows.append({
                    'Line item name': line,
                    'Note index': '',
                    'figure for 2025': None,
                    'figure for 2024': None
                })
                current_item = ""
            else:
                current_item = line
        
        i += 1
    
    return pd.DataFrame(data_rows)


def extract_financial_position_statement(pdf_file, start_page, end_page):
    """
    Extract Consolidated Statement of Financial Position from multiple pages
    """
    with pdfplumber.open(pdf_file) as pdf:
        all_lines = []
        for page_idx in range(start_page, end_page + 1):
            page = pdf.pages[page_idx]
            page_text = page.extract_text()
            lines = page_text.split('\n')
            all_lines.extend(lines)
    
    data_rows = []
    i = 0
    
    while i < len(all_lines):
        line = all_lines[i].strip()
        
        # Skip headers and metadata
        if any(skip in line for skip in [
            'Consolidated Statement of Financial Position',
            'As at 31 December',
            '2025 2024',
            'Note RMB',
            'The notes on pages',
            'consolidated financial statements',
            'approved by the board',
            'signed on its behalf',
            'Director',
            'Annual Report',
            'Ma Huateng',
            'Yang Siu Shun',
            'Tencent Holdings Limited'
        ]):
            i += 1
            continue
        
        if not line:
            i += 1
            continue
        
        # Extract numbers from the line
        numbers = re.findall(r'\(?\d+(?:,\d{3})*(?:\.\d{2,3})?\)?', line)
        
        # Check if this line has financial data
        if len(numbers) >= 2:
            val_2024_str = numbers[-1]
            val_2025_str = numbers[-2]
            
            val_2024 = convert_to_number(val_2024_str)
            val_2025 = convert_to_number(val_2025_str)
            
            # Remove numbers to get item name
            line_text = re.sub(r'\(?\d+(?:,\d{3})*(?:\.\d{2,3})?\)?', '', line).strip()
            
            data_rows.append({
                'Line item name': line_text,
                'Note index': '',
                'figure for 2025': val_2025,
                'figure for 2024': val_2024
            })
        else:
            # No financial data - could be a section header
            if line and len(line) > 2:
                if line.isupper() or line in [
                    'Non-current assets',
                    'Current assets',
                    'Total assets',
                    'Equity attributable to equity holders of the Company',
                    'Non-controlling interests',
                    'Total equity',
                    'LIABILITIES',
                    'Non-current liabilities',
                    'Current liabilities',
                    'Total liabilities',
                    'Total equity and liabilities'
                ]:
                    data_rows.append({
                        'Line item name': line,
                        'Note index': '',
                        'figure for 2025': None,
                        'figure for 2024': None
                    })
        
        i += 1
    
    return pd.DataFrame(data_rows)


def extract_cash_flows_statement(pdf_file, start_page, end_page):
    """
    Extract Consolidated Statement of Cash Flows
    """
    with pdfplumber.open(pdf_file) as pdf:
        lines = []
        for page_idx in range(start_page, end_page + 1):
            page_text = pdf.pages[page_idx].extract_text()
            lines += page_text.split('\n')
    
    data_rows = []
    buffer = []
    
    for line in lines:
        line = line.strip()
        
        if any(skip in line for skip in [
            'Consolidated Statement of Cash Flows', 'For the year ended', 'Year ended 31 December',
            '2025 2024', 'Note RMB', 'Tencent Holdings Limited', 'Annual Report', 'The notes on pages'
        ]):
            continue
        
        if not line:
            continue
        
        if line in [
            'Cash flows from operating activities',
            'Cash flows from investing activities',
            'Cash flows from financing activities'
        ]:
            data_rows.append({'Line item name': line, 'Note index': '', 'figure for 2025': None, 'figure for 2024': None})
            continue
        
        buffer.append(line)
        joined = ' '.join(buffer)
        
        # Regex patterns for matching
        dash_regex = r"[–—‑-]|None|—|\\u2013|\\u2014"
        
        # Look for two numbers at end (normal case)
        match2 = re.match(
            r"(.+?)([0-9]{1,2}\([a-z]\))?\s*(-?\(?\d[\d,\.]*\)?)\s+(-?\(?\d[\d,\.]*\)?)(?:\s*)$",
            joined)
        # Look for value + dash at end
        match1dash = re.match(
            r"(.+?)([0-9]{1,2}\([a-z]\))?\s*(-?\(?\d[\d,\.]*\)?)\s+(%s)(?:\s*)$" % dash_regex,
            joined)
        # Look for dash + value
        matchdash1 = re.match(
            r"(.+?)([0-9]{1,2}\([a-z]\))?\s+(%s)\s+(-?\(?\d[\d,\.]*\)?)(?:\s*)$" % dash_regex,
            joined)
        # One number at end only
        match1 = re.match(
            r"(.+?)([0-9]{1,2}\([a-z]\))?\s*(-?\(?\d[\d,\.]*\)?)(?:\s*)$",
            joined)
        
        if match2:
            item_text = match2.group(1).strip()
            note_index = match2.group(2).strip() if match2.group(2) else ''
            amt_2025 = convert_to_number(match2.group(3).strip())
            amt_2024 = convert_to_number(match2.group(4).strip())
            data_rows.append({'Line item name': item_text, 'Note index': note_index,
                              'figure for 2025': amt_2025, 'figure for 2024': amt_2024})
            buffer = []
        elif match1dash:
            item_text = match1dash.group(1).strip()
            note_index = match1dash.group(2).strip() if match1dash.group(2) else ''
            amt_2025 = convert_to_number(match1dash.group(3).strip())
            data_rows.append({'Line item name': item_text, 'Note index': note_index,
                              'figure for 2025': amt_2025, 'figure for 2024': None})
            buffer = []
        elif matchdash1:
            item_text = matchdash1.group(1).strip()
            note_index = matchdash1.group(2).strip() if matchdash1.group(2) else ''
            amt_2024 = convert_to_number(matchdash1.group(4).strip())
            data_rows.append({'Line item name': item_text, 'Note index': note_index,
                              'figure for 2025': None, 'figure for 2024': amt_2024})
            buffer = []
        elif match1:
            item_text = match1.group(1).strip()
            note_index = match1.group(2).strip() if match1.group(2) else ''
            amt_2025 = convert_to_number(match1.group(3).strip())
            data_rows.append({'Line item name': item_text, 'Note index': note_index,
                              'figure for 2025': amt_2025, 'figure for 2024': None})
            buffer = []
    
    return pd.DataFrame(data_rows)


def extract_changes_in_equity_statement(pdf_file, page_indices):
    """
    Extract Consolidated Statement of Changes in Equity
    """
    columns = [
        'Line item name', 'Share capital', 'Share premium', 'Treasury shares', 'Shares held for share award schemes',
        'Other reserves', 'Retained earnings', 'Total', 'Non-controlling interests', 'Total equity'
    ]
    
    with pdfplumber.open(pdf_file) as pdf:
        lines = []
        for page_index in page_indices:
            lines += pdf.pages[page_index].extract_text().split('\n')
    
    data_rows = []
    buffer = ""
    
    for line in lines:
        line = line.strip()
        
        if not line or any(skip in line for skip in [
            'Consolidated Statement of Changes in Equity', 'For the year ended', 'RMB',
            'Attributable to equity holders', 'Shares held', 'Share Share Treasury for share',
            'capital premium shares award schemes reserves earnings Total interests equity',
            'The notes on pages', 'Annual Report', 'Tencent Holdings Limited'
        ]):
            continue
        
        test_line = buffer + (" " if buffer else "") + line if buffer else line
        found_nums = re.findall(r"[\-\(0-9,.\)–]+", test_line)
        
        cleaned_nums = []
        for x in found_nums:
            if re.search(r"\d", x) or x.strip() in ['–', '-', '']:
                cleaned_nums.append(x.strip())
        
        if len(cleaned_nums) >= 8:
            # Split item name from columns
            item = test_line
            for val in cleaned_nums[-9:]:
                item = item.rsplit(val, 1)[0].strip()
            
            values = cleaned_nums[-9:]
            
            # Pad if less than 9 values
            while len(values) < 9:
                values = [''] + values
            
            rec = {'Line item name': item}
            for i, header in enumerate(columns[1:]):
                rec[header] = convert_to_number(values[i]) if values[i] not in ['', '–', '-'] else None
            
            data_rows.append(rec)
            buffer = ""
        else:
            buffer = buffer + " " + line if buffer else line
    
    return pd.DataFrame(data_rows)


# ==================== EXCEL FORMATTING ====================

def format_worksheet(worksheet, sheet_name, numeric_cols):
    """
    Apply consistent formatting to a worksheet
    """
    # Format header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format data rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for idx, cell in enumerate(row):
            cell.border = thin_border
            
            if idx == 0:  # Item name column
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif idx in numeric_cols:  # Numeric columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:  # Other columns (notes, etc.)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 60
    worksheet.row_dimensions[1].height = 30
    
    for col_idx in range(2, len(worksheet[1]) + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 18


def create_combined_excel(dataframes_dict, output_file):
    """
    Create Excel workbook with multiple sheets and formatting
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            # Determine numeric columns (all except 'Line item name' and 'Note index')
            numeric_cols = [i for i, col in enumerate(df.columns) 
                           if col not in ['Line item name', 'Note index']]
            
            format_worksheet(worksheet, sheet_name, numeric_cols)


# ==================== MAIN EXECUTION ====================

def main():
    pdf_file = "Tencent_Annual_Report_2025.pdf"
    output_excel = "Tencent_Financial_Statements_2025.xlsx"
    
    print("="*80)
    print("TENCENT 2025 - CONSOLIDATED FINANCIAL STATEMENTS EXTRACTION")
    print("="*80 + "\n")
    
    all_dataframes = {}
    extraction_failed = False
    
    # 1. INCOME STATEMENT
    print("Finding Income Statement pages...")
    try:
        income_pages = find_financial_statement_pages(
            pdf_file, 
            start_keyword="Consolidated Income Statement",
            stop_keywords=["Consolidated Statement of Comprehensive Income", "Other income"]
        )
        if not income_pages:
            raise ValueError("No pages found matching criteria")
        
        print(f"✓ Found Income Statement on pages: {income_pages}")
        df_income = extract_income_statement(pdf_file, income_pages[0])
        all_dataframes['Income Statement'] = df_income
        print(f"  Extracted {len(df_income)} rows\n")
    except Exception as e:
        print(f"✗ ERROR: Failed to extract Income Statement")
        print(f"  Reason: {str(e)}\n")
        extraction_failed = True
    
    # 2. COMPREHENSIVE INCOME STATEMENT
    print("Finding Comprehensive Income Statement pages...")
    try:
        comprehensive_pages = find_financial_statement_pages(
            pdf_file,
            start_keyword="Consolidated Statement of Comprehensive Income",
            stop_keywords=["Consolidated Statement of Financial Position"]
        )
        if not comprehensive_pages:
            raise ValueError("No pages found matching criteria")
        
        print(f"✓ Found Comprehensive Income Statement on pages: {comprehensive_pages}")
        df_comprehensive = extract_comprehensive_income_statement(pdf_file, comprehensive_pages[0])
        all_dataframes['Comprehensive Income'] = df_comprehensive
        print(f"  Extracted {len(df_comprehensive)} rows\n")
    except Exception as e:
        print(f"✗ ERROR: Failed to extract Comprehensive Income Statement")
        print(f"  Reason: {str(e)}\n")
        extraction_failed = True
    
    # 3. FINANCIAL POSITION STATEMENT
    print("Finding Financial Position Statement pages...")
    try:
        position_pages = find_financial_statement_pages(
            pdf_file,
            start_keyword="Consolidated Statement of Financial Position",
            stop_keywords=["Consolidated Statement of Changes in Equity"]
        )
        if not position_pages:
            raise ValueError("No pages found matching criteria")
        
        print(f"✓ Found Financial Position Statement on pages: {position_pages}")
        df_position = extract_financial_position_statement(pdf_file, position_pages[0], position_pages[-1])
        all_dataframes['Financial Position'] = df_position
        print(f"  Extracted {len(df_position)} rows\n")
    except Exception as e:
        print(f"✗ ERROR: Failed to extract Financial Position Statement")
        print(f"  Reason: {str(e)}\n")
        extraction_failed = True
    
    # 4. CHANGES IN EQUITY STATEMENT
    print("Finding Changes in Equity Statement pages...")
    try:
        equity_pages = find_financial_statement_pages(
            pdf_file,
            start_keyword="Consolidated Statement of Changes in Equity",
            stop_keywords=["Consolidated Statement of Cash Flows"]
        )
        if not equity_pages:
            raise ValueError("No pages found matching criteria")
        
        print(f"✓ Found Changes in Equity Statement on pages: {equity_pages}")
        df_equity = extract_changes_in_equity_statement(pdf_file, equity_pages)
        all_dataframes['Changes in Equity'] = df_equity
        print(f"  Extracted {len(df_equity)} rows\n")
    except Exception as e:
        print(f"✗ ERROR: Failed to extract Changes in Equity Statement")
        print(f"  Reason: {str(e)}\n")
        extraction_failed = True
    
    # 5. CASH FLOWS STATEMENT
    print("Finding Cash Flows Statement pages...")
    try:
        cash_pages = find_financial_statement_pages(
            pdf_file,
            start_keyword="Consolidated Statement of Cash Flows",
            stop_keywords=["Notes to the Consolidated Financial Statements"]
        )
        if not cash_pages:
            raise ValueError("No pages found matching criteria")
        
        print(f"✓ Found Cash Flows Statement on pages: {cash_pages}")
        df_cashflows = extract_cash_flows_statement(pdf_file, cash_pages[0], cash_pages[-1])
        all_dataframes['Cash Flows'] = df_cashflows
        print(f"  Extracted {len(df_cashflows)} rows\n")
    except Exception as e:
        print(f"✗ ERROR: Failed to extract Cash Flows Statement")
        print(f"  Reason: {str(e)}\n")
        extraction_failed = True
    
    # Check if any extractions failed
    if extraction_failed:
        print("\n" + "="*80)
        print("EXTRACTION INCOMPLETE")
        print("="*80)
        print("\nOne or more financial statements could not be automatically detected.")
        print("Please verify:")
        print("  1. The PDF file path is correct: " + pdf_file)
        print("  2. The PDF contains all expected financial statements")
        print("  3. The statement titles match the expected keywords")
        print("\nExiting without creating Excel file.")
        sys.exit(1)
    
    # Create combined Excel file only if all extractions succeeded
    print("Creating combined Excel file...")
    try:
        create_combined_excel(all_dataframes, output_excel)
        
        print(f"\n✓ Successfully saved all sheets to: {output_excel}")
        print("\nExtracted sheets summary:")
        for sheet_name, df in all_dataframes.items():
            print(f"  • {sheet_name}: {len(df)} rows")
        
        print("\n" + "="*80)
        print("EXTRACTION COMPLETE")
        print("="*80)
    except Exception as e:
        print(f"\n✗ ERROR: Failed to create Excel file")
        print(f"  Reason: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
